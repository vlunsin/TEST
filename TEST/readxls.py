import openpyxl

file_to_read = f"{input("file to read: ").strip()}"
print(file_to_read)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(file_to_read)

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)