# Python for automation

# Table Extraction
pip install pandas

import pandas as pd
pd.read_html(URL) # will give back a list

# Can scrape tables from website
# Can Automate downloads from website

# Reading a .csv file from a website
df_premier21 = pd.read_csv(URL)

# Show Dataframe
df_premier21 = 

# Rename columns
df_premier21.rename(columns={'Columntochange1':'Newname1',
                             'Columntochange2':'Newname2'}, inplace=True)

# Extract Tables from PDFs
# Library camelot-py tk and ghostscript
import camelot

tables = camelot.read_pdf('name.pdf', pages='1')
print(tables)
tables.export('nameofcsv', f='csv', compress=True)
tables[0]

# HTML Basics - Tags and Elements
<h1 class="title"> Titanic (1997) </h1>
<head>
<body>
<header>
<article>
<p> paragraph
<h1>, <h2>, <h3> heading

# check the site with inspect on right click to display html and check
<div> divider
<nav> navigational
<li> list item
<a> anchor  <a href="url"> text </a> or hyperlink or link
<button>
<table>
<td> table 
<tr> table row element
<iframe> page within a page, nested
<ul> unordered list
# 1 parent per node

# XPath key for webscraping with selenium 
# Similarities with CSS
//tagName
example //h1 to select all matching nodes at the same level
//tagname[1] select based on position
//tagname[@attributeName="Value"]
# Functions
contains()
starts-with()
//tagName[contains(@AttributeName,"Value")]
Can use operations
//tagName[(expression 1) and (expression 2)]
example //div/text()
//div[@class="full-script"]
# Special characters
Single / : select the children from the node set on the left side of this character
// refer to all children
. Refers to present node 
.. refers to parent node
* : wild card for all elements
@ select an Attribute 
() Grouping an Xpath expression
[n] node with index n should be selected

#Automate the news
# Chrome driver to download and after use selenium

# Driver to interact with selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

website = "URL"
path = "path of Chrome Driver"

service = Service(executable_path=path)
driver = webdriver.Chrome(service=service)
driver.get(website)
# it will open page with selenium

# Finding elements
Inspect on any website to open developer tools for html elements
Click on arrow on left and drag to element you want to check
Will give the element on the developer tool

CTRL+F on developer tool to be able to use XPATH
//div[@class='testvalue']

# use the xpath
# find element give the first element
driver.find_element(by="xpath", value='valueofxpath')

# find elements give all the elements
driver.find_elements(by="xpath", value='valueofxpath')

Open element in developer tool, to get items separately

containers = driver.find_elements(by="xpath", value='valueofxpath')

for container in containers:
    title = container.find_element(by="xpath", value='./path/path1').text # add .text to get the text from html
    subtitle = container.find_element(by="xpath", value='./path/path2').text
    link = container. find_element(by="xpath", value='./a').get_attribute("href")

    titles.append(title)
    subtit.append(subtitle)
    links.append(link)
# Links are often in href in a tag

# export to csv
# need empty lists
titles = []
subtitles = []
links = []

# Then add the list to dataframe with pandas
# useful to use dictionary, pair of key: value
my=dict = {'title':titles, 'subtitle':subtitles, 'link':links}

df_headlines = pd.DataFrame(my_dict)
df_headlines.to_csv('headline.csv')

driver.quit()

# Headless mode so that the driver doesn't open the browser
from selenium.webdriver.chrome.options import Options

options = Options()
options.headless = True
driver = webdriver.Chrome(service=service, options=options)

# Preparing script to be run everyday
# Need to convert the py file into an executable
# But need to adapt
from datetime import datetime   # for date and time
import os
import sys

application_path = os.path.dirname(sys.executable) # to get path of executable
# datetime to customize filename
now = datetime.now()
month_day_year = now.strftime("%m%d%Y") # string from time MMDDYY

# for path, messy to use / in the path
# preferably use os module os.path.join
file_name = f'headline--{month_day_year}.csv' # add f before name to format
final_path = os.path.join(application_path, file_name)
df_headlines.to_csv(final_path)  

# Convert py to exec
# py installer to convert to exec
# in terminal, need to be located where py file is, then pyinstaller --onefile nameoffile
# will create a file in dist folder

# schedule with crontab
# crontab -e schedule time with format + file path 
# exemple 0 9 * * * /filepath/file
# crontab -l for list of scheduled tasks

# Create a pivot table with Python
# first read excel files 
import pandas as pd

df = pd.read_excel('filename')

# In example, gender, product, total
df = df[['gender','product line','total']]

pivot_table = df.pivot_table(index='Gender', columns='product line', values='Total'
               aggfunc='sum' ).round(0) # round to round the numbers
pivot_table.to_excel('pivot_table.xlsx','Report, startrow=4')

# Bar chart : library openpyexcel
from openpyxl import load_workbook  # to load excel file

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# need to determine columns and rows, min and max
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

from openpyxl.chart import BarChart, Reference
barchart = BarChart()
data = Reference(sheet,
          min_col=min_column+1, # for data only
          max_col=max_column,
          min_row=min_row,  
          max_row=max_row)
categories = Reference(sheet,
          min_col=min_column, 
          max_col=min_column,
          min_row=min_row+1,   # for gender 
          max_row=max_row)

barchart.add_data(data, titles_from_data=True)  
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12")    # to add barchart

barchart.title = "TITLE"
barchart.style = 5 # number is a style, need to check separately
wb.save('barchart.xlsx')

# Excel formulas with Python
sheet['B8'] = 'SUM(B6:B7)'
sheet['B8'].style = 'Currency'

wb.save('report.xlsx')

from openpyxl.utils import get_column_letter # to get letter for column

# range you need to add for maximum value, range goes from min to max-1
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

# Format cells
from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook('file')
sheet = wb['sheetname']

sheet['A1'] = 'TITLE'
sheet['A2'] = 'Month'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

wb.save('geneatedfilename')

# Send message with Whatsapp Library pywhatkit
import pywhatkit

phone_number = input("enter : ")
# parameter to close the window after message sent : tab_close
pywhatkit.sendwhatmsg("Phonenumber","message", "Hour", "Minute", 15, True, 2)

# send message to group : need id of group
